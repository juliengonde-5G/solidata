import os
import uuid
from datetime import datetime
from openpyxl import load_workbook
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession
from app.models.collecte import CAV, RouteTemplate, RouteTemplatePoint, WeightHistory, Vehicle


async def run_import(filetype: str, db: AsyncSession):
    """Import Excel files (CAV, tournées, pesées)."""
    results = {"cav": 0, "routes": 0, "pesees": 0, "vehicules": 0, "errors": []}

    upload_dir = "/app/uploads"

    if filetype in ("all", "tonnage"):
        # Chercher le fichier tonnage
        for fname in os.listdir(upload_dir):
            if "tonnage" in fname.lower() and fname.endswith(".xlsx"):
                try:
                    count = await import_tonnage(os.path.join(upload_dir, fname), db)
                    results["pesees"] = count
                except Exception as e:
                    results["errors"].append(f"Tonnage: {str(e)}")
                break

    if filetype in ("all", "tournees"):
        for fname in os.listdir(upload_dir):
            if "tournee" in fname.lower() and fname.endswith(".xlsx"):
                try:
                    cav_count, route_count = await import_tournees(os.path.join(upload_dir, fname), db)
                    results["cav"] = cav_count
                    results["routes"] = route_count
                except Exception as e:
                    results["errors"].append(f"Tournées: {str(e)}")
                break

    return results


async def import_tonnage(filepath: str, db: AsyncSession):
    """Import pesées historiques depuis Excel."""
    wb = load_workbook(filepath, read_only=True)
    ws = wb.active
    count = 0
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    for row in rows:
        if not row or not row[0]:
            continue
        try:
            wh = WeightHistory(
                external_id=str(row[0]) if row[0] else None,
                origine=str(row[1]) if len(row) > 1 and row[1] else None,
                categorie=str(row[2]) if len(row) > 2 and row[2] else None,
                poids_net=float(row[3]) if len(row) > 3 and row[3] else None,
                tare=float(row[4]) if len(row) > 4 and row[4] else None,
                poids_brut=float(row[5]) if len(row) > 5 and row[5] else None,
                date_pesee=row[6] if len(row) > 6 and isinstance(row[6], datetime) else None,
            )
            if wh.date_pesee:
                wh.mois = wh.date_pesee.month
                wh.trimestre = (wh.date_pesee.month - 1) // 3 + 1
                wh.annee = wh.date_pesee.year
            db.add(wh)
            count += 1
        except Exception:
            continue
    await db.commit()
    return count


async def import_tournees(filepath: str, db: AsyncSession):
    """Import CAV et tournées depuis Excel."""
    wb = load_workbook(filepath, read_only=True)
    ws = wb.active
    cav_count = 0
    route_count = 0

    # Identifier les colonnes
    headers = [cell for cell in next(ws.iter_rows(min_row=1, max_row=1, values_only=True))]

    routes_cache = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not row[0]:
            continue
        try:
            # Créer/récupérer le CAV
            nom = str(row[0]).strip()
            existing = await db.execute(select(CAV).where(CAV.nom == nom))
            cav = existing.scalar_one_or_none()
            if not cav:
                cav = CAV(
                    nom=nom,
                    adresse=str(row[1]).strip() if len(row) > 1 and row[1] else None,
                    code_postal=str(row[2]).strip() if len(row) > 2 and row[2] else None,
                    ville=str(row[3]).strip() if len(row) > 3 and row[3] else None,
                    latitude=float(row[4]) if len(row) > 4 and row[4] else None,
                    longitude=float(row[5]) if len(row) > 5 and row[5] else None,
                    nb_cav=int(row[6]) if len(row) > 6 and row[6] else 1,
                    frequence_passage=int(row[7]) if len(row) > 7 and row[7] else 1,
                    qr_code=f"ST-{cav_count + 1:04d}-{uuid.uuid4().hex[:6].upper()}",
                )
                db.add(cav)
                await db.flush()
                cav_count += 1

            # Tournée
            tournee_nom = str(row[8]).strip() if len(row) > 8 and row[8] else None
            if tournee_nom:
                if tournee_nom not in routes_cache:
                    existing_rt = await db.execute(select(RouteTemplate).where(RouteTemplate.nom == tournee_nom))
                    rt = existing_rt.scalar_one_or_none()
                    if not rt:
                        rt = RouteTemplate(nom=tournee_nom)
                        db.add(rt)
                        await db.flush()
                        route_count += 1
                    routes_cache[tournee_nom] = rt.id

                ordre = int(row[9]) if len(row) > 9 and row[9] else 1
                existing_rtp = await db.execute(
                    select(RouteTemplatePoint)
                    .where(RouteTemplatePoint.route_template_id == routes_cache[tournee_nom])
                    .where(RouteTemplatePoint.cav_id == cav.id)
                )
                if not existing_rtp.scalar_one_or_none():
                    rtp = RouteTemplatePoint(
                        route_template_id=routes_cache[tournee_nom],
                        cav_id=cav.id,
                        ordre=ordre,
                    )
                    db.add(rtp)
        except Exception:
            continue

    await db.commit()
    return cav_count, route_count
